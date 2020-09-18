import openpyxl

class TestExcelDemo:
    def testreadExcelData(self):
        print("This is test data utility")
        book = openpyxl.load_workbook("C:\\Users\\Public\\Automation\\E2EAutomationData.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        print(cell)
        sheet.cell(row=2, column=2).value = "Anuj"
        print(sheet.max_column)
        print(sheet.max_row)
